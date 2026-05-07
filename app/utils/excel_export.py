"""
app/utils/excel_export.py
Экспорт кандидатов в Excel
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any

from app.core.logger import get_logger

logger = get_logger(__name__)


def export_candidates_to_excel(candidates: List[Dict[str, Any]], filename: str = None) -> str:
    """Экспортирует список кандидатов в Excel файл"""
    if filename is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"candidates_export_{timestamp}.xlsx"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Кандидаты"
    
    headers = ["ID", "ФИО", "Email", "Телефон", "Последняя должность", "Опыт", "Последняя Компания", "Навыки", "Дата создания"]
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for cand in candidates:
        skills = cand.get('skills', [])
        skills_str = ', '.join(skills[:10]) + ('...' if len(skills) > 10 else '')
        
        created_at = cand.get('created_at')
        if created_at:
            if isinstance(created_at, datetime):
                created_at_str = created_at.strftime('%Y-%m-%d')
            else:
                created_at_str = str(created_at)[:10] if str(created_at) else '—'
        else:
            created_at_str = '—'
        
        row = [
            cand.get('id', ''),
            cand.get('name', '—'),
            cand.get('email', '—'),
            cand.get('phone', '—'),
            cand.get('last_position', '—'),
            cand.get('experience_years', 0),
            cand.get('last_company', '—'),
            skills_str,
            created_at_str
        ]
        ws.append(row)
    
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    filepath = Path(filename)
    wb.save(filepath)
    logger.info(f"✅ Экспортировано {len(candidates)} кандидатов в {filepath}")
    
    return str(filepath)
